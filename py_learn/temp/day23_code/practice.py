'''
将多个内容一致的Excel文件合并到一个Excel文件中

小宝剑大药房（高新店）2018年销售数据.xlsx
小宝剑大药房（犀浦店）2018年销售数据.xlsx
小宝剑大药房（新津店）2018年销售数据.xlsx

'''
import openpyxl
total_info_list1 = []

def read_excel(path:str):
    wb_read1 = openpyxl.load_workbook(path)
    sheet1 = wb_read1.active
    print(sheet1)
    print(sheet1.dimensions)
    headers = [cell.value for cell in sheet1[2]]
    print(headers)
    item_list = []
    for i in range(3, sheet1.max_row + 1):
        temp_row = sheet1[i]
        item = [cell.value for cell in temp_row if cell.value != None]
        print(item)
        if item:
            item_list.append(item)
    return headers, item_list

def write_excel(headers,info_list:list):
    wb = openpyxl.Workbook()
    wb.active.title = '小宝剑大药房汇总'
    sheet1 = wb['小宝剑大药房汇总']
    sheet1.append(['小宝剑大药房2018年汇总数据'])
    sheet1.append(headers)
    for li in info_list:
        sheet1.append(li)
    wb.save(r'data/小宝剑大药房各店汇总数据.xlsx')

def extend_total_info(path):
    headers,list_temp = read_excel(path)
    total_info_list1.extend(list_temp)
    return headers

if __name__ == '__main__':
    extend_total_info(r'data/小宝剑大药房（新津店）2018年销售数据.xlsx')
    extend_total_info(r'data/小宝剑大药房（犀浦店）2018年销售数据.xlsx')
    headers = extend_total_info(r'data/小宝剑大药房（高新店）2018年销售数据.xlsx')
    write_excel(headers, total_info_list1)



'''
import openpyxl

wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = '小宝剑大药房汇总'
sheet2.append(['小宝剑大药房2018年汇总数据'])
sheet2.append(['购药时间','社保卡号','商品编码','商品名称','销售数量','应收金额','实收金额'])
shop_lst = []
def getvalue(xlname):
    wb = openpyxl.load_workbook(f'小宝剑大药房（{xlname}）2018年销售数据.xlsx')
    sheet = wb.active
    for n in range(3,sheet.max_row + 1):
        row_list = []
        row = sheet[n]
        for cell in row:
            if cell.value != None:
                row_list.append(cell.value)
        if row_list != []:
            shop_lst.append(row_list)
getvalue('新津店')
getvalue('犀浦店')
getvalue('高新店')
for sublist in shop_lst:
    sheet2.append(sublist)

wb2.save(r'小宝剑大药房汇总.xlsx')
'''

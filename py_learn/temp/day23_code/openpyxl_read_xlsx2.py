# openpyxl 读取xlsx
import openpyxl
# 1.获取工作簿对象
wb =openpyxl.load_workbook(r'data/阿里巴巴2020年股票数据.xlsx')
print(wb)

# 2.获取工作表对象
# a.获取工作表名称
print(wb.sheetnames)        # ['股票数据']
# b.获取工作表对象
sheet1 = wb.active  # active :活跃，获取工作簿中活跃的工作表对象
print(sheet1)
# c.wb[工作表名称]
sheet2 = wb['股票数据']
print(sheet2)
# d. wb.worksheets  :获取一个列表，其中的元素是工作表对象
#   wb.worksheets[索引]： 获取知道索引处的工作表对象
sheet3 = wb.worksheets[0]
print(sheet3)

# 3.获取制定工作表的行数和列数【范围】
print(sheet3.max_row, sheet3.max_column)        # 255 7
print(sheet3.dimensions)        # A1:G255

# 4.获取单元格对象
# a. sheet['A3']
cell1 = sheet3['A1']
print(cell1)
# b. sheet.cell(row,col)
'''
openpyxl:
    row:
        Excel: 1 2 3 4 .....
        Python:1 2 3 4 .....
    col:
        Excel:A B C D....
        Python:1 2 3 4 .....
'''

cell2 = sheet3.cell(1,1)    # A1
print(cell2)

# c.获取单元格的值
print(cell1.value)
print(cell2.value)

# 5.获取指定行，指定列
# 获取指定行
row = sheet3[1]
print(row)
# for cell in row:
#     print(cell.value)

# 获取指定列
col = sheet3['A']
print(col)
# for cell in col:
#     print(cell.value)

# 获取多行多列
cells = sheet3['A2:C8']
print(cells)
for row in cells:       # 获取每行
    for cell in row:    # 获取每一行中的列
        print(cell.value)

# 练习：获取student-score.xlsx 中成绩表中有哪些小学，总共有几个小学
wb = openpyxl.load_workbook(r'data/student-score.xlsx')
sheet = wb['成绩表']
c_col = sheet['C'][2:]
print(c_col)
school = set([cell.value for cell in c_col])
print(school)
print(len(school))

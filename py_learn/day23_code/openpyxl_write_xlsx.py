import openpyxl
# openpyxl 写入数据
# 注意：当写入数据时候，一定要保证目标文件处于关闭状态，否则会报权限被拒绝的错误



# 1.创建工作簿
# 注意：当我们创建一个工作簿是，默认会创建一张工作表
wb = openpyxl.Workbook()
print(wb.sheetnames)    # ['Sheet']
sheet = wb['Sheet']
sheet.title = '学生表'
print(wb.sheetnames)    # ['学生表']


#2.创建和删除工作表【了解】
wb.create_sheet()
print(wb.sheetnames)    # ['学生表', 'Sheet']
wb.create_sheet()
print(wb.sheetnames)        # ['学生表', 'Sheet', 'Sheet1']

# 复制表
sheet = wb['学生表']
wb.copy_worksheet(sheet)
print(wb.sheetnames)

# 删除表
sheet1 = wb['Sheet1']
wb.remove(sheet1)
print(wb.sheetnames)


# 3.修改或者追加单元格
sheet = wb['学生表']
# a.给单元格添加数据
# 方式一
sheet['A1'] = '姓名'
sheet['B1'] = '性别'
sheet['C1'] = '成绩'

# 方式二
# sheet.cell(row,col,value)
sheet.cell(1,4,'爱好')     # D1

# b.追加一行数据
# sheet.append(['张三','male',89,'跳舞'])

# c.追加多行数据
# 注意：无论数据量多大，如果想要将数据写入到excel中，只需要将数据整理成以下形式
# 二维列表，二维元组，元组嵌套列表，列表嵌套元组
student_list = [
['张四','male',77,'跳舞'],
['张五','male',67,'跳舞'],
['张六','male',54,'跳舞'],
['张七','male',79,'打羽毛球'],
['张八','male',59,'打篮球'],
['张九','male',93,'玩游戏']
]
for stu in student_list:
    sheet.append(stu)

# 保存工作簿
wb.save(r'data/test01.xlsx')

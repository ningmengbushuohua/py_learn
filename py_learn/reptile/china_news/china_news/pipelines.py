# Define your item pipelines here
# 管道
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import openpyxl


class ChinaNewsPipeline:
    def process_item(self, item, spider):
        return item

class ChinaNewsToExcel:
    '''
    将数据存储到excel 文件中
    '''
    def open_spider(self,spider):
        # 命令行执行爬虫时，方法触发
        self.workbook = openpyxl.Workbook()
        col_name = ['新闻类型','新闻标题','新闻链接','新闻时间']
        self.workbook.active.title = '中国新闻网'
        self.workbook.active.append(col_name)

    def close_spider(self,spider):
        # 命令行关闭爬虫时，方法触发
        self.workbook.save(r'data.xlsx')

    def process_item(self, item, spider):
        '''
        数据处理过程在此完成
        :param item:
        :param spider:
        :return:
        '''
        data_list = [item['news_type'],item['news_title'],item['news_link'],item['news_time']]
        # 方法一：
        self.workbook.active.append(data_list)
        # 方法二：
        # 每次执行都会计算当前有多少行
        # row_index = self.workbook['中国新闻网'].max_row
        # col_index = 1
        # for i in range(len(data_list)):
        #     self.workbook.active.cell(row_index + 1,col_index).value = data_list[i]
        #     col_index += 1
        return item
